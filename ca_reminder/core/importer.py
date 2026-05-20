"""
Import clients and compliances from CSV or XLSX files.
Also generates blank templates users can fill in.

CSV / XLSX column definitions
──────────────────────────────────────────────────────────────
Clients:
  client_code | name* | email* | phone | pan | gstin | consent_given | notes

Compliances:
  name* | description | category* | frequency* | due_day | due_month | advance_reminder_days

Fields marked * are required; blank rows are skipped gracefully.
"""

import csv
import io
import logging
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

# ── Column definitions ────────────────────────────────────────────────────────

CLIENT_COLUMNS = [
    "client_code", "name", "email", "phone", "pan", "gstin",
    "consent_given", "notes",
]
CLIENT_REQUIRED = {"name", "email"}

CLIENT_SAMPLE_ROWS = [
    ["C001", "Ramesh Kumar",  "ramesh@example.com",  "9876543210", "ABCDE1234F", "27ABCDE1234F1Z5", "yes", "Pvt Ltd company"],
    ["C002", "Priya Sharma",  "priya@example.com",   "9123456789", "FGHIJ5678K", "",                "yes", "Proprietor – retail"],
    ["C003", "Vikram Nair",   "vikram@example.com",  "",           "LMNOP9012Q", "32LMNOP9012Q1Z1", "yes", "Partnership firm"],
]

COMPLIANCE_COLUMNS = [
    "name", "description", "category", "frequency",
    "due_day", "due_month", "advance_reminder_days",
]
COMPLIANCE_REQUIRED = {"name", "category", "frequency"}

COMPLIANCE_SAMPLE_ROWS = [
    ["GSTR-1 (Monthly)", "Monthly outward supply return",  "GST",       "Monthly",   "11", "",   "7"],
    ["TDS Payment",      "Monthly TDS deposit",            "TDS / TCS", "Monthly",   "7",  "",   "3"],
    ["ITR – Individual", "Annual income-tax return",       "Income Tax","Annual",    "31", "7",  "30"],
    ["Advance Tax Q1",   "15 June instalment",             "Income Tax","Annual",    "15", "6",  "7"],
]

VALID_CATEGORIES = {
    "GST", "TDS / TCS", "Income Tax", "ROC / MCA",
    "PF / ESI", "Professional Tax", "FEMA / RBI", "Custom / Others",
}
VALID_FREQUENCIES = {"Monthly", "Quarterly", "Half-Yearly", "Annual", "Event-Based"}

import calendar as _cal
MONTH_MAP = {m.lower(): i for i, m in enumerate(_cal.month_name) if m}
MONTH_MAP.update({_cal.month_abbr[i].lower(): i for i in range(1, 13)})


# ── Public: read file ─────────────────────────────────────────────────────────

def read_clients(filepath: str) -> tuple[list[dict], list[str]]:
    """
    Parse a CSV or XLSX file into a list of client dicts.
    Returns (rows, warnings).
    """
    raw = _read_file(filepath)
    rows, warnings = [], []
    for i, row in enumerate(raw, start=2):
        row = {k.strip().lower(): str(v).strip() for k, v in row.items()}
        # normalise keys (allow spaces / mixed case in file)
        row = _normalise_keys(row, CLIENT_COLUMNS)
        missing = CLIENT_REQUIRED - {k for k, v in row.items() if v}
        if missing:
            warnings.append(f"Row {i}: skipped — missing {', '.join(missing)}")
            continue
        client = {
            "client_code":   row.get("client_code", ""),
            "name":          row.get("name", ""),
            "email":         row.get("email", ""),
            "phone":         row.get("phone", ""),
            "pan":           row.get("pan", "").upper(),
            "gstin":         row.get("gstin", "").upper(),
            "consent_given": _truthy(row.get("consent_given", "no")),
            "notes":         row.get("notes", ""),
        }
        rows.append(client)
    return rows, warnings


def read_compliances(filepath: str) -> tuple[list[dict], list[str]]:
    """
    Parse a CSV or XLSX file into a list of compliance dicts.
    Returns (rows, warnings).
    """
    raw = _read_file(filepath)
    rows, warnings = [], []
    for i, row in enumerate(raw, start=2):
        row = {k.strip().lower(): str(v).strip() for k, v in row.items()}
        row = _normalise_keys(row, COMPLIANCE_COLUMNS)
        missing = COMPLIANCE_REQUIRED - {k for k, v in row.items() if v}
        if missing:
            warnings.append(f"Row {i}: skipped — missing {', '.join(missing)}")
            continue

        # due_day
        due_day = None
        if row.get("due_day"):
            try:
                due_day = int(row["due_day"])
                if not 1 <= due_day <= 31:
                    raise ValueError
            except ValueError:
                warnings.append(f"Row {i}: invalid due_day '{row['due_day']}' — ignored")

        # due_month (accept number or name)
        due_month = None
        dm_raw = row.get("due_month", "")
        if dm_raw:
            try:
                dm = int(dm_raw)
                if 1 <= dm <= 12:
                    due_month = dm
            except ValueError:
                due_month = MONTH_MAP.get(dm_raw.lower())
            if due_month is None:
                warnings.append(f"Row {i}: unrecognised due_month '{dm_raw}' — ignored")

        # advance days
        adv = 7
        try:
            adv = int(row.get("advance_reminder_days", "7") or "7")
        except ValueError:
            pass

        comp = {
            "name":                  row.get("name", ""),
            "description":           row.get("description", ""),
            "category":              _best_match(row.get("category", ""), VALID_CATEGORIES),
            "frequency":             _best_match(row.get("frequency", ""), VALID_FREQUENCIES),
            "due_day":               due_day,
            "due_month":             due_month,
            "advance_reminder_days": adv,
        }
        rows.append(comp)
    return rows, warnings


# ── Public: template generators ──────────────────────────────────────────────

def save_client_template_csv(filepath: str) -> None:
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(CLIENT_COLUMNS)
        w.writerows(CLIENT_SAMPLE_ROWS)


def save_compliance_template_csv(filepath: str) -> None:
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(COMPLIANCE_COLUMNS)
        w.writerows(COMPLIANCE_SAMPLE_ROWS)


def save_client_template_xlsx(filepath: str) -> None:
    openpyxl = _require_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients"
    _xlsx_header(ws, CLIENT_COLUMNS)
    for row in CLIENT_SAMPLE_ROWS:
        ws.append(row)
    _xlsx_notes(ws, {
        1: "Optional — your internal client code",
        2: "Required — client's full name",
        3: "Required — email address for reminders",
        4: "Phone number (optional)",
        5: "PAN number (optional)",
        6: "GSTIN (optional)",
        7: "yes / no  — must be 'yes' to receive reminders (DPDP Act consent)",
        8: "Any additional notes",
    })
    _xlsx_autofit(ws)
    wb.save(filepath)


def save_compliance_template_xlsx(filepath: str) -> None:
    openpyxl = _require_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compliances"
    _xlsx_header(ws, COMPLIANCE_COLUMNS)
    for row in COMPLIANCE_SAMPLE_ROWS:
        ws.append(row)
    _xlsx_notes(ws, {
        1: "Required — compliance name",
        2: "Optional description",
        3: f"Required — one of: {', '.join(sorted(VALID_CATEGORIES))}",
        4: f"Required — one of: {', '.join(sorted(VALID_FREQUENCIES))}",
        5: "Day of month (1-31) when compliance is due",
        6: "Month number (1-12) or name — for Annual frequency",
        7: "Days before due date to include in reminder (default 7)",
    })
    _xlsx_autofit(ws)
    wb.save(filepath)


def openpyxl_available() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


# ── Internal helpers ──────────────────────────────────────────────────────────

def _read_file(filepath: str) -> list[dict]:
    path = Path(filepath)
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx(filepath)
    if suffix in (".csv", ".txt"):
        return _read_csv(filepath)
    raise ValueError(f"Unsupported file type: {suffix}  (use .csv or .xlsx)")


def _read_csv(filepath: str) -> list[dict]:
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [dict(row) for row in reader]


def _read_xlsx(filepath: str) -> list[dict]:
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    result = []
    for row in rows[1:]:
        if all(c is None for c in row):
            continue
        result.append({headers[i]: (str(row[i]).strip() if row[i] is not None else "")
                       for i in range(len(headers))})
    return result


def _require_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX support.\n"
            "Install it with:  pip install openpyxl"
        )


def _truthy(val: str) -> bool:
    return str(val).strip().lower() in ("yes", "y", "true", "1", "✓", "x")


def _normalise_keys(row: dict, columns: list[str]) -> dict:
    """Map header variations → canonical column names."""
    out = {}
    for col in columns:
        # exact match first
        if col in row:
            out[col] = row[col]
            continue
        # fuzzy: strip spaces, lowercase
        col_norm = col.replace("_", " ").replace("-", " ")
        for k, v in row.items():
            if k.replace("_", " ").replace("-", " ").lower() == col_norm:
                out[col] = v
                break
        else:
            out[col] = ""
    return out


def _best_match(value: str, valid_set: set) -> str:
    """Return exact match or best case-insensitive match; original value if no match."""
    if value in valid_set:
        return value
    for v in valid_set:
        if v.lower() == value.lower():
            return v
    return value


def _xlsx_header(ws, columns: list) -> None:
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        ws.append(columns)
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="1D1D1F")
            cell.alignment = Alignment(horizontal="left")
    except Exception:
        ws.append(columns)


def _xlsx_notes(ws, notes: dict) -> None:
    """Add a Notes row at the bottom of the sheet."""
    try:
        from openpyxl.styles import Font
        ws.append([])
        ws.append(["# Column notes:"])
        ws["A" + str(ws.max_row)].font = Font(italic=True, color="888888")
        for col_idx, note in notes.items():
            ws.append([f"  Column {col_idx} ({ws.cell(1, col_idx).value}): {note}"])
            ws.cell(ws.max_row, 1).font = Font(italic=True, color="888888")
    except Exception:
        pass


def _xlsx_autofit(ws) -> None:
    try:
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    except Exception:
        pass
