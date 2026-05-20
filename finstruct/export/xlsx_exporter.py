"""XLSX export using openpyxl — formatted Excel workbook."""

from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from ..core.fs_engine import FSDocument, FSLine
from ..core.notes_engine import Note

P_BLUE   = "FF0078D4"
P_DARK   = "FF106EBE"
P_LIGHT  = "FFC7E0F4"
P_ALT    = "FFEFF6FC"
P_WHITE  = "FFFFFFFF"
P_TOTAL  = "FF106EBE"
P_GRAND  = "FF0078D4"
P_WARN   = "FFFCE4E4"

THIN = Side(style="thin", color="FFEDEBE9")
THICK = Side(style="medium", color="FF0078D4")
NUM_FMT = '#,##0.00'


def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)


def _font(bold=False, white=False, sz=9) -> Font:
    return Font(name="Segoe UI", bold=bold, size=sz,
                color="FFFFFFFF" if white else "FF201F1E")


def _border() -> Border:
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_fs_sheet(ws, lines: list[FSLine], sheet_title: str, em: dict, fy: str):
    entity_name = em.get("entity_name", em.get("Company_Name", "Entity"))
    ws.append([entity_name.upper()])
    ws["A1"].font = Font(name="Segoe UI", bold=True, size=13, color="FF0078D4")
    ws.merge_cells("A1:D1")
    ws.append([sheet_title])
    ws["A2"].font = Font(name="Segoe UI", bold=True, size=11)
    ws.merge_cells("A2:D2")
    ws.append([f"FY {fy}  |  All amounts in ₹ unless otherwise stated"])
    ws["A3"].font = Font(name="Segoe UI", size=8, italic=True, color="FF605E5C")
    ws.merge_cells("A3:D3")
    ws.append([])

    hdr_row = 5
    ws.append(["Particulars", "Note", "Current Year ₹", "Previous Year ₹"])
    for col, w in zip(["A","B","C","D"], [55, 8, 18, 18]):
        ws.column_dimensions[col].width = w
    for c in range(1, 5):
        cell = ws.cell(hdr_row, c)
        cell.fill    = _fill(P_BLUE)
        cell.font    = _font(bold=True, white=True)
        cell.border  = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[hdr_row].height = 22

    r = hdr_row + 1
    for i, ln in enumerate(lines):
        if ln.row_type == "BLANK":
            ws.append(["", "", "", ""])
            r += 1; continue

        indent = "    " * ln.indent
        cy_val = ln.cy if ln.row_type not in ("SECTION","HEADER","BLANK","TEXT") else None
        py_val = ln.py if ln.row_type not in ("SECTION","HEADER","BLANK","TEXT") else None
        note_v = ln.note if ln.note else ""

        ws.append([indent + ln.label, note_v, cy_val, py_val])

        a_cell = ws.cell(r, 1)
        b_cell = ws.cell(r, 2)
        c_cell = ws.cell(r, 3)
        d_cell = ws.cell(r, 4)

        for cell in [a_cell, b_cell, c_cell, d_cell]:
            cell.border = _border()

        if ln.row_type == "HEADER":
            for c in range(1, 5):
                ws.cell(r, c).fill = _fill(P_BLUE)
                ws.cell(r, c).font = _font(bold=True, white=True)
            ws.merge_cells(f"A{r}:D{r}")
        elif ln.row_type == "SECTION":
            for c in range(1, 5):
                ws.cell(r, c).fill = _fill(P_LIGHT)
                ws.cell(r, c).font = Font(name="Segoe UI", bold=True, size=9,
                                          color="FF003087")
            ws.merge_cells(f"A{r}:D{r}")
        elif ln.row_type == "GRAND":
            for c in range(1, 5):
                ws.cell(r, c).fill = _fill(P_GRAND)
                ws.cell(r, c).font = _font(bold=True, white=True)
        elif ln.row_type == "TOTAL":
            for c in range(1, 5):
                ws.cell(r, c).fill = _fill(P_TOTAL)
                ws.cell(r, c).font = _font(bold=True, white=True)
        elif ln.row_type == "SUBTOTAL":
            a_cell.font = _font(bold=True)
        elif i % 2 == 0:
            for c in range(1, 5):
                ws.cell(r, c).fill = _fill(P_ALT)

        if cy_val is not None:
            c_cell.number_format = NUM_FMT
            c_cell.alignment = Alignment(horizontal="right")
        if py_val is not None:
            d_cell.number_format = NUM_FMT
            d_cell.alignment = Alignment(horizontal="right")
        b_cell.alignment = Alignment(horizontal="center")
        r += 1

    ws.freeze_panes = f"A{hdr_row+1}"


def _write_note_sheet(ws, note: Note):
    ws.append([f"Note {note.number}: {note.title}"])
    ws["A1"].font = Font(name="Segoe UI", bold=True, size=11, color="FF0078D4")
    ws.merge_cells("A1:C1")
    ws.append([])
    ws.append(["Particulars", "Current Year ₹", "Previous Year ₹"])
    for col, w in zip(["A","B","C"], [55, 18, 18]):
        ws.column_dimensions[col].width = w
    for c in range(1, 4):
        cell = ws.cell(3, c)
        cell.fill   = _fill(P_BLUE)
        cell.font   = _font(bold=True, white=True)
        cell.border = _border()

    r = 4
    for i, ln in enumerate(note.lines):
        if ln.row_type == "BLANK":
            ws.append(["","",""]); r += 1; continue
        indent = "    " * ln.indent
        cy_v = ln.cy if ln.row_type not in ("SECTION","HEADER","BLANK","TEXT") else None
        py_v = ln.py if ln.row_type not in ("SECTION","HEADER","BLANK","TEXT") else None
        ws.append([indent + ln.label, cy_v, py_v])
        for c in range(1, 4):
            ws.cell(r, c).border = _border()
        if ln.row_type in ("TOTAL","GRAND"):
            for c in range(1, 4):
                ws.cell(r, c).fill = _fill(P_TOTAL)
                ws.cell(r, c).font = _font(bold=True, white=True)
        elif ln.row_type == "SECTION":
            for c in range(1, 4):
                ws.cell(r, c).fill = _fill(P_LIGHT)
                ws.cell(r, c).font = Font(name="Segoe UI", bold=True, size=9, color="FF003087")
            ws.merge_cells(f"A{r}:C{r}")
        elif i % 2 == 0:
            for c in range(1, 4):
                ws.cell(r, c).fill = _fill(P_ALT)
        if cy_v is not None:
            ws.cell(r, 2).number_format = NUM_FMT
            ws.cell(r, 2).alignment = Alignment(horizontal="right")
        if py_v is not None:
            ws.cell(r, 3).number_format = NUM_FMT
            ws.cell(r, 3).alignment = Alignment(horizontal="right")
        r += 1
    ws.freeze_panes = "A4"


def export_xlsx(doc: FSDocument, notes: list[Note], output_path: Path):
    wb = Workbook()
    em = doc.entity_master

    sheet_map = [
        ("BS",  doc.bs,  "Balance Sheet"),
        ("PL",  doc.pl,  "Profit & Loss"),
        ("IE",  doc.ie,  "Income & Expenditure"),
        ("RP",  doc.rp,  "Receipt & Payment"),
        ("CF",  doc.cf,  "Cash Flow"),
    ]
    first = True
    for tab, lines, title in sheet_map:
        if not lines:
            continue
        if first:
            ws = wb.active
            ws.title = tab
            first = False
        else:
            ws = wb.create_sheet(tab)
        _write_fs_sheet(ws, lines, title, em, doc.fy)

    for note in notes:
        ws = wb.create_sheet(f"N{note.number}")
        _write_note_sheet(ws, note)

    wb.save(output_path)
